"""
Travel Tienda scraper.

Uses the public Oracle Commerce endpoints exposed by https://tienda.travel.cl:
- /ccstore/v1/search for keyword/general listings
- /ccstore/v1/products for category/product data

The module keeps requests small and paginated, then flattens products for
preview and Excel export.
"""
from __future__ import annotations

import concurrent.futures
import json
import math
import re
import time
from io import BytesIO
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlencode, urljoin, unquote
from urllib.request import Request, urlopen


BASE_URL = "https://tienda.travel.cl"
SITE_ID = "tienda"
PRICE_GROUP = "tienda"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)
DEFAULT_PAGE_SIZE = 100
MAX_WORKERS = 5
REQUEST_TIMEOUT = 25
REQUEST_RETRIES = 3
CACHE_TTL = 300

_CACHE: dict[str, tuple[float, Any]] = {}


def _cache_get(key: str) -> Any | None:
    entry = _CACHE.get(key)
    if not entry:
        return None
    expires_at, value = entry
    if expires_at < time.time():
        _CACHE.pop(key, None)
        return None
    return value


def _cache_set(key: str, value: Any) -> None:
    _CACHE[key] = (time.time() + CACHE_TTL, value)


def _request(url: str, accept: str = "application/json") -> bytes:
    cached = _cache_get(url)
    if cached is not None:
        return cached

    headers = {
        "User-Agent": USER_AGENT,
        "Accept": accept,
        "X-CCSite": SITE_ID,
        "X-CCPriceListGroup": PRICE_GROUP,
    }
    last_exc: BaseException | None = None
    for attempt in range(REQUEST_RETRIES):
        try:
            req = Request(url, headers=headers)
            with urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
                body = resp.read()
                _cache_set(url, body)
                return body
        except (HTTPError, URLError, TimeoutError, OSError) as exc:
            last_exc = exc
            if isinstance(exc, HTTPError) and exc.code in (400, 403, 404):
                raise
            time.sleep(min(1.0 * (2**attempt), 6.0))
    if last_exc:
        raise last_exc
    raise RuntimeError("No se pudo leer tienda.travel.cl")


def _fetch_json(url: str) -> Any:
    return json.loads(_request(url).decode("utf-8", errors="ignore"))


def _fetch_text(url: str) -> str:
    return _request(url, accept="text/html").decode("utf-8", errors="ignore")


def _first(attrs: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = attrs.get(key)
        if isinstance(value, list) and value:
            return value[0]
        if value not in (None, ""):
            return value
    return None


def _num(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        raw = str(value).strip().replace("$", "").replace(" ", "")
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        return float(raw)
    except (TypeError, ValueError):
        return 0.0


def _money(value: Any) -> str:
    number = _num(value)
    if number <= 0:
        return ""
    return f"$ {int(round(number)):,}".replace(",", ".")


def _absolute_url(path: str | None) -> str:
    if not path:
        return ""
    return urljoin(BASE_URL, path)


def _discount(normal: float, offer: float) -> int | None:
    if normal > 0 and offer > 0 and normal > offer:
        return round(((normal - offer) / normal) * 100)
    return None


def fetch_home_state() -> dict[str, Any]:
    html = _fetch_text(f"{BASE_URL}/")
    match = re.search(r'window\.state = JSON\.parse\(decodeURI\("(.*?)"\)\)', html)
    if not match:
        raise RuntimeError("No se encontro el estado inicial de Travel Tienda.")
    return json.loads(unquote(match.group(1)))


def _category_maps() -> tuple[dict[str, dict[str, Any]], list[str]]:
    state = fetch_home_state()
    categories = state.get("catalogRepository", {}).get("categories", {}) or {}
    roots = (
        state.get("clientRepository", {})
        .get("context", {})
        .get("global", {})
        .get("categories", [])
        or []
    )
    return categories, roots


def fetch_categories() -> list[dict[str, Any]]:
    categories, roots = _category_maps()
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()

    def walk(category_id: str, path: list[str], depth: int) -> None:
        if not category_id or category_id in seen:
            return
        cat = categories.get(category_id) or {}
        name = cat.get("displayName") or category_id
        current_path = [*path, name]
        rows.append({
            "id": category_id,
            "name": name,
            "path": " / ".join(current_path),
            "route": cat.get("route") or "",
            "depth": depth,
        })
        seen.add(category_id)
        for child in cat.get("childCategories") or []:
            walk(child.get("id"), current_path, depth + 1)

    for root_id in roots:
        walk(root_id, [], 0)

    for category_id in categories:
        walk(category_id, [], 0)

    return rows


def _descendant_category_ids(category_id: str) -> list[str]:
    categories, _roots = _category_maps()
    ids: list[str] = []
    seen: set[str] = set()

    def walk(cid: str) -> None:
        if not cid or cid in seen:
            return
        seen.add(cid)
        ids.append(cid)
        for child in (categories.get(cid) or {}).get("childCategories") or []:
            walk(child.get("id"))

    walk(category_id)
    return ids


def _category_name(category_id: str | None) -> str:
    if not category_id:
        return ""
    categories, _roots = _category_maps()
    return (categories.get(category_id) or {}).get("displayName") or category_id


def _build_search_url(query: str = "", offset: int = 0, limit: int = DEFAULT_PAGE_SIZE) -> str:
    params: dict[str, Any] = {"No": max(0, offset), "Nrpp": max(1, limit)}
    if query.strip():
        params["Ntt"] = query.strip()
    return f"{BASE_URL}/ccstore/v1/search?{urlencode(params)}"


def _build_category_url(category_id: str, offset: int = 0, limit: int = DEFAULT_PAGE_SIZE) -> str:
    params = {
        "categoryId": category_id,
        "limit": max(1, limit),
        "offset": max(0, offset),
        "withPrices": "true",
    }
    return f"{BASE_URL}/ccstore/v1/products?{urlencode(params)}"


def search_page(query: str = "", offset: int = 0, limit: int = DEFAULT_PAGE_SIZE) -> dict[str, Any]:
    return _fetch_json(_build_search_url(query, offset, limit))


def category_page(category_id: str, offset: int = 0, limit: int = DEFAULT_PAGE_SIZE) -> dict[str, Any]:
    return _fetch_json(_build_category_url(category_id, offset, limit))


def _normalize_search_record(record: dict[str, Any], position: int) -> dict[str, Any]:
    top_attrs = record.get("attributes") or {}
    child_attrs = {}
    child_records = record.get("records") or []
    if child_records:
        child_attrs = child_records[0].get("attributes") or {}
    attrs = {**top_attrs, **child_attrs}

    sku = _first(attrs, "product.repositoryId", "sku.repositoryId") or _first(child_attrs, "sku.repositoryId")
    name = _first(attrs, "product.displayName", "sku.displayName") or _first(child_attrs, "sku.displayName") or ""
    brand = _first(attrs, "product.brand") or ""
    category = _first(attrs, "parentCategory.displayName", "product.category") or ""
    route = _first(attrs, "product.route") or ""
    image = _first(attrs, "product.primaryMediumImageURL", "product.primaryFullImageURL", "product.primarySmallImageURL")

    normal_raw = _num(_first(child_attrs, "sku.listPrice") or _first(attrs, "sku.listPrice", "product.listPrice"))
    offer_raw = _num(
        _first(child_attrs, "sku.salePrice", "sku.activePrice")
        or _first(attrs, "sku.salePrice", "sku.activePrice", "product.salePrice")
    )
    if offer_raw <= 0:
        offer_raw = normal_raw

    return {
        "position": position,
        "id": sku or "",
        "name": name,
        "brand": brand,
        "category": category,
        "normal_price": _money(normal_raw),
        "offer_price": _money(offer_raw),
        "normal_price_raw": normal_raw,
        "offer_price_raw": offer_raw,
        "discount_percent": _discount(normal_raw, offer_raw),
        "picture": _absolute_url(image),
        "link": _absolute_url(route),
    }


def _normalize_product(item: dict[str, Any], position: int, fallback_category_id: str | None = None) -> dict[str, Any]:
    sku = item.get("id") or item.get("repositoryId") or ""
    normal_raw = _num((item.get("listPrices") or {}).get(PRICE_GROUP) or item.get("listPrice"))
    offer_raw = _num((item.get("salePrices") or {}).get(PRICE_GROUP) or item.get("salePrice") or normal_raw)
    category_id = fallback_category_id
    if not category_id:
        parent = item.get("parentCategory") or item.get("defaultParentCategory") or {}
        category_id = parent.get("repositoryId") if isinstance(parent, dict) else ""

    return {
        "position": position,
        "id": sku,
        "name": item.get("displayName") or item.get("productDisplayName") or "",
        "brand": item.get("brand") or "",
        "category": _category_name(category_id),
        "normal_price": _money(normal_raw),
        "offer_price": _money(offer_raw),
        "normal_price_raw": normal_raw,
        "offer_price_raw": offer_raw,
        "discount_percent": _discount(normal_raw, offer_raw),
        "picture": _absolute_url(item.get("primaryMediumImageURL") or item.get("primaryFullImageURL")),
        "link": _absolute_url(item.get("route") or (f"/product/{quote(str(sku))}" if sku else "")),
    }


def _search_records(data: dict[str, Any]) -> tuple[list[dict[str, Any]], int]:
    results = data.get("resultsList") or data.get("results") or {}
    return results.get("records") or [], int(results.get("totalNumRecs") or 0)


def _collect_search(query: str, limit: int, fetch_all: bool) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    page_size = DEFAULT_PAGE_SIZE if fetch_all else min(DEFAULT_PAGE_SIZE, max(1, limit))
    first = search_page(query, 0, page_size)
    records, total = _search_records(first)
    target = min(total, 10000 if fetch_all else limit)

    pages = max(1, math.ceil(target / page_size))
    page_map: dict[int, list[dict[str, Any]]] = {0: records}

    def fetch(offset: int) -> tuple[int, list[dict[str, Any]]]:
        data = search_page(query, offset, page_size)
        recs, _total = _search_records(data)
        return offset, recs

    offsets = [page * page_size for page in range(1, pages)]
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(fetch, offset): offset for offset in offsets}
        for future in concurrent.futures.as_completed(futures):
            offset, recs = future.result()
            page_map[offset] = recs

    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    for offset in sorted(page_map):
        for record in page_map[offset]:
            item = _normalize_search_record(record, len(items) + 1)
            sku = item.get("id")
            if sku and sku in seen:
                continue
            if sku:
                seen.add(sku)
            items.append(item)
            if len(items) >= target:
                break
        if len(items) >= target:
            break

    return items, {"total": total, "pages_fetched": len(page_map), "fetched_raw": sum(len(v) for v in page_map.values())}


def _collect_category(category_id: str, query: str, limit: int, fetch_all: bool) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    category_ids = _descendant_category_ids(category_id) or [category_id]
    page_size = DEFAULT_PAGE_SIZE
    target = min(10000 if fetch_all else limit, 10000)
    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    total = 0
    pages_fetched = 0
    fetched_raw = 0

    for cid in category_ids:
        if len(items) >= target:
            break
        first = category_page(cid, 0, min(page_size, max(1, target - len(items))))
        cat_total = int(first.get("totalResults") or 0)
        total += cat_total
        pages_fetched += 1
        fetched_raw += len(first.get("items") or [])
        category_items = {0: first.get("items") or []}
        pages = max(1, math.ceil(min(cat_total, target - len(items)) / page_size))

        def fetch(offset: int) -> tuple[int, list[dict[str, Any]]]:
            data = category_page(cid, offset, page_size)
            return offset, data.get("items") or []

        offsets = [page * page_size for page in range(1, pages)]
        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(fetch, offset): offset for offset in offsets}
            for future in concurrent.futures.as_completed(futures):
                offset, raw_items = future.result()
                category_items[offset] = raw_items
                pages_fetched += 1
                fetched_raw += len(raw_items)

        for offset in sorted(category_items):
            for raw in category_items[offset]:
                item = _normalize_product(raw, len(items) + 1, cid)
                sku = item.get("id")
                if sku and sku in seen:
                    continue
                if sku:
                    seen.add(sku)
                if query and query.lower() not in item.get("name", "").lower():
                    continue
                items.append(item)
                if len(items) >= target:
                    break
            if len(items) >= target:
                break

    return items, {"total": total, "pages_fetched": pages_fetched, "fetched_raw": fetched_raw}


def apply_filters(
    items: list[dict[str, Any]],
    min_price: float = 0,
    max_price: float = 0,
    include_words: list[str] | None = None,
    exclude_words: list[str] | None = None,
    ordering: str = "relevance",
) -> list[dict[str, Any]]:
    filtered = []
    include_words = include_words or []
    exclude_words = exclude_words or []
    for item in items:
        price = item.get("offer_price_raw") or item.get("normal_price_raw") or 0
        if min_price > 0 and price < min_price:
            continue
        if max_price > 0 and price > max_price:
            continue
        name = (item.get("name") or "").lower()
        if include_words and not all(word.lower() in name for word in include_words if word.strip()):
            continue
        if exclude_words and any(word.lower() in name for word in exclude_words if word.strip()):
            continue
        filtered.append(item)

    if ordering == "price_asc":
        filtered.sort(key=lambda x: x.get("offer_price_raw") or 0)
    elif ordering == "price_desc":
        filtered.sort(key=lambda x: x.get("offer_price_raw") or 0, reverse=True)
    elif ordering == "discount_desc":
        filtered.sort(key=lambda x: x.get("discount_percent") or 0, reverse=True)
    elif ordering == "name_asc":
        filtered.sort(key=lambda x: (x.get("name") or "").lower())

    for idx, item in enumerate(filtered, start=1):
        item["position"] = idx
    return filtered


def collect_results(
    query: str = "",
    category_id: str = "",
    ordering: str = "relevance",
    limit: int = 200,
    fetch_all: bool = False,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    cleaned_query = query.strip()
    if category_id:
        try:
            items, meta = _collect_category(category_id, cleaned_query, limit, fetch_all)
        except Exception as exc:
            if not cleaned_query:
                raise
            items, meta = _collect_search(cleaned_query, limit, fetch_all)
            meta["category_fallback"] = True
            meta["category_error"] = str(exc)
        else:
            if cleaned_query and not items:
                fallback_items, fallback_meta = _collect_search(cleaned_query, limit, fetch_all)
                if fallback_items:
                    items = fallback_items
                    meta = {
                        **fallback_meta,
                        "category_fallback": True,
                        "category_empty": True,
                        "requested_category_id": category_id,
                    }
    else:
        items, meta = _collect_search(cleaned_query, limit, fetch_all)
    items = apply_filters(items, ordering=ordering)
    return items[: min(limit, 10000)], meta


def count_products(query: str = "", category_id: str = "") -> dict[str, Any]:
    if category_id:
        total = 0
        for cid in _descendant_category_ids(category_id) or [category_id]:
            data = category_page(cid, 0, 1)
            total += int(data.get("totalResults") or 0)
        return {"count": total, "count_source": "travel_products_api"}
    data = search_page(query.strip(), 0, 1)
    _records, total = _search_records(data)
    return {"count": total, "count_source": "travel_search_api"}


def build_xlsx_bytes(items: list[dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Travel Export"
    headers = ["Posicion", "SKU", "Nombre", "Marca", "Categoria", "Precio Normal", "Precio Oferta", "Descuento %", "Link"]
    fill = PatternFill(start_color="0A1861", end_color="0A1861", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    for row, item in enumerate(items, start=2):
        ws.cell(row=row, column=1, value=item.get("position", row - 1))
        ws.cell(row=row, column=2, value=item.get("id", ""))
        ws.cell(row=row, column=3, value=item.get("name", ""))
        ws.cell(row=row, column=4, value=item.get("brand", ""))
        ws.cell(row=row, column=5, value=item.get("category", ""))
        ws.cell(row=row, column=6, value=item.get("normal_price", ""))
        ws.cell(row=row, column=7, value=item.get("offer_price", ""))
        discount = item.get("discount_percent")
        ws.cell(row=row, column=8, value=f"{discount}%" if discount else "")
        ws.cell(row=row, column=9, value=item.get("link", ""))

    widths = [10, 16, 62, 18, 24, 18, 18, 12, 64]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
