"""
SoloTodo.cl API Client — scraper module.

Uses the public API at https://publicapi.solotodo.com/ to fetch
categories, products (with prices from multiple stores), and export
results to Excel.  Designed for large-list extraction (up to 10 000+
products per export).
"""
from __future__ import annotations

import concurrent.futures
import json
import math
import sys
import time
from io import BytesIO
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse, urlencode
from urllib.request import Request, urlopen

BASE_URL = "https://publicapi.solotodo.com"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)
DEFAULT_PAGE_SIZE = 100
MAX_WORKERS = 6
REQUEST_RETRIES = 3
REQUEST_TIMEOUT = 25

# ---------------------------------------------------------------------------
# In-memory page cache
# ---------------------------------------------------------------------------
PAGE_CACHE_TTL = 300
_PAGE_CACHE: dict[str, tuple[float, Any]] = {}


def _cache_get(key: str) -> Any | None:
    entry = _PAGE_CACHE.get(key)
    if not entry:
        return None
    expires_at, value = entry
    if expires_at < time.time():
        _PAGE_CACHE.pop(key, None)
        return None
    return value


def _cache_set(key: str, value: Any) -> None:
    _PAGE_CACHE[key] = (time.time() + PAGE_CACHE_TTL, value)


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------
def _fetch_json(url: str, timeout: int = REQUEST_TIMEOUT) -> Any:
    cached = _cache_get(url)
    if cached is not None:
        return cached

    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "application/json",
    }
    last_exc: BaseException | None = None
    for attempt in range(REQUEST_RETRIES):
        try:
            req = Request(url, headers=headers)
            with urlopen(req, timeout=timeout) as resp:
                data = json.loads(resp.read().decode("utf-8", errors="ignore"))
                _cache_set(url, data)
                return data
        except (HTTPError, URLError, TimeoutError, OSError) as exc:
            last_exc = exc
            if isinstance(exc, HTTPError) and exc.code in (400, 403, 404):
                raise
            time.sleep(min(1.0 * (2 ** attempt), 8.0))
    if last_exc:
        raise last_exc
    raise RuntimeError("No se pudo leer la API de SoloTodo.")


def _progress(prefix: str, current: int, total: int | None = None) -> None:
    if total and total > 0:
        pct = min(100, int((current / total) * 100))
        bar_w = 24
        filled = int((pct / 100) * bar_w)
        bar = "#" * filled + "-" * (bar_w - filled)
        msg = f"\r{prefix} [{bar}] {pct:3d}% ({current}/{total})"
    else:
        msg = f"\r{prefix} {current}"
    print(msg, end="", file=sys.stderr, flush=True)


def _progress_done() -> None:
    print(file=sys.stderr, flush=True)


# ---------------------------------------------------------------------------
# Categories
# ---------------------------------------------------------------------------
def fetch_categories() -> list[dict[str, Any]]:
    """Return list of {id, name, slug} for every SoloTodo category."""
    url = f"{BASE_URL}/categories/?format=json&page_size=500"
    data = _fetch_json(url)
    if isinstance(data, list):
        raw = data
    elif isinstance(data, dict):
        raw = data.get("results") or data.get("data") or []
        if not raw and isinstance(data, list):
            raw = data
    else:
        raw = []

    cats = []
    for item in raw:
        cats.append({
            "id": item.get("id"),
            "name": item.get("name") or "",
            "slug": item.get("slug") or "",
        })
    cats.sort(key=lambda c: c["name"])
    return cats


# ---------------------------------------------------------------------------
# Products: browse (with prices)
# ---------------------------------------------------------------------------
def _build_browse_url(
    query: str = "",
    category_id: int | None = None,
    country_id: int = 1,  # 1 = Chile
    ordering: str = "offer_price_usd",
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE,
) -> str:
    params: dict[str, Any] = {
        "format": "json",
        "page_size": page_size,
        "page": page,
        "countries": country_id,
        "ordering": ordering,
    }
    if query.strip():
        params["search"] = query.strip()
    if category_id:
        params["categories"] = category_id
    return f"{BASE_URL}/products/browse/?{urlencode(params)}"


def browse_products_page(
    query: str = "",
    category_id: int | None = None,
    country_id: int = 1,
    ordering: str = "offer_price_usd",
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE,
) -> dict[str, Any]:
    url = _build_browse_url(query, category_id, country_id, ordering, page, page_size)
    return _fetch_json(url)


def _iter_browse_entries(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Return product entries from either old flat results or current bucketed API."""
    entries: list[dict[str, Any]] = []
    for raw in results:
        product_entries = raw.get("product_entries")
        if isinstance(product_entries, list):
            entries.extend([entry for entry in product_entries if isinstance(entry, dict)])
        elif raw.get("product") or raw.get("metadata"):
            entries.append(raw)
    return entries


def _format_clp(value: Any) -> str:
    try:
        amount = int(round(float(value)))
    except (TypeError, ValueError):
        return ""
    return f"$ {amount:,}".replace(",", ".")


def _parse_price(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _extract_category_id(category_url: str) -> int | None:
    if not category_url:
        return None
    parts = urlparse(category_url).path.rstrip("/").split("/")
    try:
        return int(parts[-1])
    except (TypeError, ValueError):
        return None


def _normalize_browse_result(raw: dict[str, Any], position: int) -> dict[str, Any]:
    """Flatten a browse API result into a simple dict for export."""
    product = raw.get("product") or {}
    metadata = raw.get("metadata") or {}
    specs = product.get("specs") or {}

    prices = metadata.get("prices_per_currency") or []
    normal_price = ""
    offer_price = ""
    normal_price_raw = 0.0
    offer_price_raw = 0.0
    if prices:
        p = prices[0]
        normal_price_raw = _parse_price(p.get("normal_price"))
        offer_price_raw = _parse_price(p.get("offer_price"))
        normal_price = _format_clp(normal_price_raw)
        offer_price = _format_clp(offer_price_raw)

    discount_pct = None
    if normal_price_raw > 0 and offer_price_raw > 0 and normal_price_raw > offer_price_raw:
        discount_pct = round(((normal_price_raw - offer_price_raw) / normal_price_raw) * 100)

    brand = specs.get("brand_unicode") or specs.get("family_brand_unicode") or ""
    name = product.get("name") or specs.get("unicode") or ""
    category_url = product.get("category") or ""
    picture = product.get("picture_url") or ""
    slug = product.get("slug") or ""
    product_id = product.get("id") or 0
    category_id = _extract_category_id(category_url)

    return {
        "position": position,
        "id": product_id,
        "name": name,
        "brand": brand,
        "normal_price": normal_price,
        "offer_price": offer_price,
        "normal_price_raw": normal_price_raw,
        "offer_price_raw": offer_price_raw,
        "discount_percent": discount_pct,
        "picture": picture,
        "slug": slug,
        "link": f"https://www.solotodo.cl/products/{product_id}-{slug}" if product_id else "",
        "category_url": category_url,
        "category_id": category_id,
    }


def collect_browse_results(
    query: str = "",
    category_id: int | None = None,
    country_id: int = 1,
    ordering: str = "offer_price_usd",
    limit: int = 200,
    fetch_all: bool = False,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """
    Fetch browse results with automatic pagination.
    Returns (items, meta) where meta has total, pages_fetched, etc.
    """
    page_size = min(DEFAULT_PAGE_SIZE, limit) if not fetch_all else DEFAULT_PAGE_SIZE

    # First page
    first = browse_products_page(query, category_id, country_id, ordering, 1, page_size)
    total = first.get("count") or 0
    results_raw = _iter_browse_entries(first.get("results") or [])

    target = min(limit, total)
    total_pages = max(1, math.ceil(target / page_size))

    items: list[dict[str, Any]] = []
    seen: set[int] = set()

    for raw in results_raw:
        product = raw.get("product") or {}
        pid = product.get("id")
        if pid and pid in seen:
            continue
        if pid:
            seen.add(pid)
        item = _normalize_browse_result(raw, len(items) + 1)
        items.append(item)
        if len(items) >= target:
            break

    pages_fetched = 1
    fetched_raw = len(results_raw)

    if len(items) < target and total_pages > 1:
        remaining_pages = list(range(2, total_pages + 1))

        def fetch_page(page_num: int) -> tuple[int, list[dict]]:
            data = browse_products_page(query, category_id, country_id, ordering, page_num, page_size)
            return page_num, _iter_browse_entries(data.get("results") or [])

        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(fetch_page, p): p for p in remaining_pages}
            page_results: dict[int, list[dict]] = {}
            for future in concurrent.futures.as_completed(futures):
                try:
                    page_num, page_data = future.result()
                    page_results[page_num] = page_data
                except HTTPError as exc:
                    if exc.code == 404:
                        continue
                    raise

        for page_num in sorted(page_results.keys()):
            results = page_results[page_num]
            fetched_raw += len(results)
            pages_fetched += 1
            _progress("Scraping SoloTodo", len(items), target)
            for raw in results:
                product = raw.get("product") or {}
                pid = product.get("id")
                if pid and pid in seen:
                    continue
                if pid:
                    seen.add(pid)
                item = _normalize_browse_result(raw, len(items) + 1)
                items.append(item)
                if len(items) >= target:
                    break
            if len(items) >= target:
                break

    _progress_done()

    meta = {
        "total": total,
        "pages_fetched": pages_fetched,
        "fetched_raw": fetched_raw,
        "target": target,
    }
    return items[:target], meta


# ---------------------------------------------------------------------------
# Filters
# ---------------------------------------------------------------------------
def apply_filters(
    items: list[dict[str, Any]],
    min_price: float = 0,
    max_price: float = 0,
    include_words: list[str] | None = None,
    exclude_words: list[str] | None = None,
) -> list[dict[str, Any]]:
    filtered = []
    for item in items:
        price = item.get("offer_price_raw") or item.get("normal_price_raw") or 0

        if min_price > 0 and price < min_price:
            continue
        if max_price > 0 and price > max_price:
            continue

        name_lower = (item.get("name") or "").lower()

        if include_words:
            if not all(w.lower() in name_lower for w in include_words if w.strip()):
                continue

        if exclude_words:
            if any(w.lower() in name_lower for w in exclude_words if w.strip()):
                continue

        filtered.append(item)

    # Re-index positions
    for idx, item in enumerate(filtered, start=1):
        item["position"] = idx

    return filtered


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------
def build_xlsx_bytes(items: list[dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "SoloTodo Export"

    headers = [
        "Posicion", "Nombre", "Marca", "Precio Normal",
        "Precio Oferta", "Descuento %", "Link",
    ]
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(items, start=2):
        ws.cell(row=row_idx, column=1, value=item.get("position", row_idx - 1))
        ws.cell(row=row_idx, column=2, value=item.get("name", ""))
        ws.cell(row=row_idx, column=3, value=item.get("brand", ""))
        ws.cell(row=row_idx, column=4, value=item.get("normal_price", ""))
        ws.cell(row=row_idx, column=5, value=item.get("offer_price", ""))

        discount = item.get("discount_percent")
        ws.cell(row=row_idx, column=6, value=f"{discount}%" if discount else "")

        ws.cell(row=row_idx, column=7, value=item.get("link", ""))

    # Auto-size columns (approximate)
    col_widths = [10, 60, 18, 18, 18, 12, 55]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Convenience: count only (fast, 1 page)
# ---------------------------------------------------------------------------
def count_products(
    query: str = "",
    category_id: int | None = None,
    country_id: int = 1,
    ordering: str = "offer_price_usd",
) -> dict[str, Any]:
    data = browse_products_page(query, category_id, country_id, ordering, 1, 1)
    total = data.get("count") or 0
    return {
        "count": total,
        "count_source": "solotodo_api",
    }
